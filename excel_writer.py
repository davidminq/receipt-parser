"""
Excel export module for receipt data.

This module provides comprehensive Excel export functionality with data preservation,
backup management, and multiple format support following SOLID principles.
"""

import shutil
from abc import ABC, abstractmethod
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Dict, Any, Union
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config import settings
from models import Receipt, ReceiptItem, ExportRequest, ExportFormat, ProcessingError, ValidationError
from utils.logging_setup import LoggerMixin, log_function_call


class ExcelWriterInterface(ABC):
    """Abstract interface for Excel writing implementations."""

    @abstractmethod
    def save_receipt(self, receipt: Receipt, file_path: Path) -> None:
        """Save single receipt to Excel file.

        Args:
            receipt: Receipt data to save
            file_path: Path to Excel file
        """
        pass

    @abstractmethod
    def save_receipts(self, receipts: List[Receipt], file_path: Path) -> None:
        """Save multiple receipts to Excel file.

        Args:
            receipts: List of receipts to save
            file_path: Path to Excel file
        """
        pass


class BackupManager(LoggerMixin):
    """Manages backup operations for Excel files."""

    def __init__(self):
        """Initialize backup manager."""
        self.max_backups = settings.excel.max_backup_files

    def create_backup(self, file_path: Path) -> Optional[Path]:
        """Create backup of existing file.

        Args:
            file_path: Path to file to backup

        Returns:
            Path to backup file or None if backup not needed
        """
        if not file_path.exists():
            return None

        if not settings.excel.create_backup:
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"
        backup_path = file_path.parent / backup_name

        try:
            shutil.copy2(file_path, backup_path)
            self.logger.info("Backup created", original=str(file_path), backup=str(backup_path))

            # Clean old backups
            self._cleanup_old_backups(file_path)

            return backup_path

        except Exception as e:
            self.logger.error("Failed to create backup", file=str(file_path), error=str(e))
            return None

    def _cleanup_old_backups(self, original_file: Path) -> None:
        """Clean up old backup files."""
        if self.max_backups <= 0:
            return

        backup_pattern = f"{original_file.stem}_backup_*{original_file.suffix}"
        backup_files = list(original_file.parent.glob(backup_pattern))

        if len(backup_files) <= self.max_backups:
            return

        # Sort by modification time and keep only the newest
        backup_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        files_to_delete = backup_files[self.max_backups:]

        for backup_file in files_to_delete:
            try:
                backup_file.unlink()
                self.logger.debug("Old backup deleted", file=str(backup_file))
            except Exception as e:
                self.logger.warning("Failed to delete old backup", file=str(backup_file), error=str(e))


class DataFrameBuilder(LoggerMixin):
    """Builds pandas DataFrames from receipt data."""

    def __init__(self):
        """Initialize DataFrame builder."""
        self.column_config = settings.excel.columns

    def build_receipt_dataframe(self, receipt: Receipt) -> pd.DataFrame:
        """Build DataFrame from single receipt.

        Args:
            receipt: Receipt data

        Returns:
            Formatted DataFrame
        """
        if not receipt.items:
            return pd.DataFrame(columns=list(self.column_config.values()))

        # Build data rows
        data_rows = []
        for item in receipt.items:
            row = {
                self.column_config["date"]: receipt.date.strftime("%Y-%m-%d"),
                self.column_config["vendor"]: item.vendor,
                self.column_config["category"]: item.category or "",
                self.column_config["amount"]: float(item.amount)
            }
            data_rows.append(row)

        df = pd.DataFrame(data_rows)

        # Add total row
        total_row = {
            self.column_config["date"]: receipt.date.strftime("%Y-%m-%d"),
            self.column_config["vendor"]: "TOTAL",
            self.column_config["category"]: "",
            self.column_config["amount"]: receipt.total_float
        }
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        return df

    def build_summary_dataframe(self, receipts: List[Receipt]) -> pd.DataFrame:
        """Build summary DataFrame from multiple receipts.

        Args:
            receipts: List of receipts

        Returns:
            Summary DataFrame
        """
        if not receipts:
            return pd.DataFrame()

        summary_data = []
        for receipt in receipts:
            summary_row = {
                "Date": receipt.date.strftime("%Y-%m-%d"),
                "Items": receipt.item_count,
                "Total": receipt.total_float,
                "Source": str(receipt.source_file) if receipt.source_file else "Manual"
            }
            summary_data.append(summary_row)

        return pd.DataFrame(summary_data)


class ExcelFormatter(LoggerMixin):
    """Handles Excel formatting and styling."""

    def __init__(self):
        """Initialize Excel formatter."""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.currency_format = settings.excel.currency_format
        self.date_format = settings.excel.date_format

    def format_worksheet(self, worksheet, df: pd.DataFrame) -> None:
        """Apply formatting to worksheet.

        Args:
            worksheet: Excel worksheet
            df: DataFrame with data
        """
        if df.empty:
            return

        # Format headers
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")

        # Format amount column
        amount_col = self._find_amount_column(df)
        if amount_col:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{amount_col}{row}"]
                cell.number_format = self.currency_format

        # Format date column
        date_col = self._find_date_column(df)
        if date_col:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{date_col}{row}"]
                cell.number_format = self.date_format

        # Adjust column widths
        self._auto_adjust_columns(worksheet)

        # Add borders
        self._add_borders(worksheet)

    def _find_amount_column(self, df: pd.DataFrame) -> Optional[str]:
        """Find the amount column letter."""
        amount_col_name = settings.excel.columns["amount"]
        if amount_col_name in df.columns:
            col_index = df.columns.get_loc(amount_col_name) + 1
            return chr(64 + col_index)  # Convert to Excel column letter
        return None

    def _find_date_column(self, df: pd.DataFrame) -> Optional[str]:
        """Find the date column letter."""
        date_col_name = settings.excel.columns["date"]
        if date_col_name in df.columns:
            col_index = df.columns.get_loc(date_col_name) + 1
            return chr(64 + col_index)  # Convert to Excel column letter
        return None

    def _auto_adjust_columns(self, worksheet) -> None:
        """Auto-adjust column widths."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_borders(self, worksheet) -> None:
        """Add borders to data range."""
        if worksheet.max_row < 2:
            return

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                     min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border


class ExcelFileManager(LoggerMixin):
    """Manages Excel file operations with data preservation."""

    def __init__(self, backup_manager: Optional[BackupManager] = None,
                 formatter: Optional[ExcelFormatter] = None):
        """Initialize Excel file manager.

        Args:
            backup_manager: Backup manager instance
            formatter: Excel formatter instance
        """
        self.backup_manager = backup_manager or BackupManager()
        self.formatter = formatter or ExcelFormatter()

    def save_dataframe_safely(self, df: pd.DataFrame, file_path: Path, sheet_name: str) -> None:
        """Save DataFrame to Excel file while preserving existing data.

        Args:
            df: DataFrame to save
            file_path: Path to Excel file
            sheet_name: Name of the sheet

        Raises:
            ProcessingError: If save operation fails
        """
        try:
            # Ensure parent directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            # Create backup if file exists
            self.backup_manager.create_backup(file_path)

            # Save to Excel
            if file_path.exists():
                self._append_to_existing_file(df, file_path, sheet_name)
            else:
                self._create_new_file(df, file_path, sheet_name)

            self.logger.info("Excel file saved successfully", file=str(file_path), sheet=sheet_name)

        except Exception as e:
            self.logger.error("Failed to save Excel file", file=str(file_path), error=str(e))
            raise ProcessingError(f"Failed to save Excel file: {str(e)}")

    def _append_to_existing_file(self, df: pd.DataFrame, file_path: Path, sheet_name: str) -> None:
        """Append DataFrame to existing Excel file."""
        # Load existing workbook
        workbook = load_workbook(file_path)

        # Replace sheet if it exists, otherwise add new sheet
        if sheet_name in workbook.sheetnames:
            # Remove existing sheet
            workbook.remove(workbook[sheet_name])

        # Create new sheet
        worksheet = workbook.create_sheet(sheet_name)

        # Add data
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        # Apply formatting
        self.formatter.format_worksheet(worksheet, df)

        # Save workbook
        workbook.save(file_path)
        workbook.close()

    def _create_new_file(self, df: pd.DataFrame, file_path: Path, sheet_name: str) -> None:
        """Create new Excel file with DataFrame."""
        workbook = Workbook()

        # Remove default sheet and create new one
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(sheet_name)

        # Add data
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        # Apply formatting
        self.formatter.format_worksheet(worksheet, df)

        # Save workbook
        workbook.save(file_path)
        workbook.close()

    def read_existing_receipts(self, file_path: Path) -> List[Receipt]:
        """Read existing receipts from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of existing receipts
        """
        if not file_path.exists():
            return []

        try:
            # Read all sheets
            excel_data = pd.read_excel(file_path, sheet_name=None)
            receipts = []

            for sheet_name, df in excel_data.items():
                if sheet_name.startswith("Receipt "):
                    receipt = self._dataframe_to_receipt(df, sheet_name)
                    if receipt:
                        receipts.append(receipt)

            return receipts

        except Exception as e:
            self.logger.warning("Failed to read existing receipts", file=str(file_path), error=str(e))
            return []

    def _dataframe_to_receipt(self, df: pd.DataFrame, sheet_name: str) -> Optional[Receipt]:
        """Convert DataFrame back to Receipt object."""
        if df.empty:
            return None

        try:
            # Extract date from sheet name
            date_str = sheet_name.replace("Receipt ", "")
            receipt_date = datetime.strptime(date_str, "%Y-%m-%d").date()

            # Convert rows to ReceiptItems (excluding total row)
            items = []
            for _, row in df.iterrows():
                vendor = row.get(settings.excel.columns["vendor"], "")
                if vendor == "TOTAL":
                    continue

                amount = row.get(settings.excel.columns["amount"], 0)
                category = row.get(settings.excel.columns["category"], "")

                item = ReceiptItem(
                    vendor=vendor,
                    amount=Decimal(str(amount)),
                    category=category
                )
                items.append(item)

            return Receipt(date=receipt_date, items=items)

        except Exception as e:
            self.logger.warning("Failed to convert DataFrame to Receipt", sheet=sheet_name, error=str(e))
            return None


class ReceiptExcelWriter(ExcelWriterInterface, LoggerMixin):
    """Main Excel writer implementation following SOLID principles."""

    def __init__(self,
                 dataframe_builder: Optional[DataFrameBuilder] = None,
                 file_manager: Optional[ExcelFileManager] = None):
        """Initialize Excel writer.

        Args:
            dataframe_builder: DataFrame builder instance
            file_manager: File manager instance
        """
        self.dataframe_builder = dataframe_builder or DataFrameBuilder()
        self.file_manager = file_manager or ExcelFileManager()

    @log_function_call
    def save_receipt(self, receipt: Receipt, file_path: Path) -> None:
        """Save single receipt to Excel file.

        Args:
            receipt: Receipt data to save
            file_path: Path to Excel file

        Raises:
            ValidationError: If receipt data is invalid
            ProcessingError: If save operation fails
        """
        self._validate_receipt(receipt)
        self._validate_file_path(file_path)

        # Build DataFrame
        df = self.dataframe_builder.build_receipt_dataframe(receipt)

        # Generate sheet name
        sheet_name = settings.excel.sheet_name_format.format(date=receipt.date.strftime("%Y-%m-%d"))

        # Save to file
        self.file_manager.save_dataframe_safely(df, file_path, sheet_name)

        self.logger.info(
            "Receipt saved to Excel",
            file=str(file_path),
            sheet=sheet_name,
            items=receipt.item_count,
            total=receipt.total_float
        )

    @log_function_call
    def save_receipts(self, receipts: List[Receipt], file_path: Path) -> None:
        """Save multiple receipts to Excel file.

        Args:
            receipts: List of receipts to save
            file_path: Path to Excel file

        Raises:
            ValidationError: If receipt data is invalid
            ProcessingError: If save operation fails
        """
        if not receipts:
            raise ValidationError("Cannot save empty receipt list", "receipts")

        self._validate_file_path(file_path)

        # Save each receipt to its own sheet
        for receipt in receipts:
            self.save_receipt(receipt, file_path)

        # Create summary sheet
        summary_df = self.dataframe_builder.build_summary_dataframe(receipts)
        if not summary_df.empty:
            self.file_manager.save_dataframe_safely(summary_df, file_path, "Summary")

        self.logger.info(
            "Multiple receipts saved to Excel",
            file=str(file_path),
            receipt_count=len(receipts),
            total_items=sum(r.item_count for r in receipts)
        )

    def _validate_receipt(self, receipt: Receipt) -> None:
        """Validate receipt data."""
        if not isinstance(receipt, Receipt):
            raise ValidationError(f"Expected Receipt object, got {type(receipt)}", "receipt_type")

        if not receipt.items:
            raise ValidationError("Receipt has no items", "receipt_items")

        if not isinstance(receipt.date, date):
            raise ValidationError("Receipt date must be a date object", "receipt_date")

    def _validate_file_path(self, file_path: Path) -> None:
        """Validate file path."""
        if not isinstance(file_path, Path):
            raise ValidationError(f"Expected Path object, got {type(file_path)}", "file_path_type")

        if file_path.suffix.lower() not in {".xlsx", ".xls"}:
            raise ValidationError(f"Invalid Excel file extension: {file_path.suffix}", "file_extension")

        # Check if parent directory is writable
        parent_dir = file_path.parent
        if parent_dir.exists() and not parent_dir.is_dir():
            raise ValidationError(f"Parent path is not a directory: {parent_dir}", "parent_directory")


class ExcelExportManager(LoggerMixin):
    """High-level Excel export management."""

    def __init__(self, excel_writer: Optional[ExcelWriterInterface] = None):
        """Initialize export manager.

        Args:
            excel_writer: Excel writer implementation
        """
        self.excel_writer = excel_writer or ReceiptExcelWriter()

    def export_receipts(self, export_request: ExportRequest) -> None:
        """Export receipts based on request.

        Args:
            export_request: Export configuration

        Raises:
            ProcessingError: If export fails
        """
        if export_request.format != ExportFormat.EXCEL:
            raise ProcessingError(f"Unsupported export format: {export_request.format}")

        try:
            if len(export_request.receipts) == 1:
                self.excel_writer.save_receipt(export_request.receipts[0], export_request.output_path)
            else:
                self.excel_writer.save_receipts(export_request.receipts, export_request.output_path)

        except (ValidationError, ProcessingError):
            raise
        except Exception as e:
            raise ProcessingError(f"Unexpected export error: {str(e)}")


# Backward compatibility function
def save_to_excel(data: List[tuple], receipt_date: date, file_path: Optional[str] = None) -> None:
    """Legacy function for backward compatibility.

    Args:
        data: List of (vendor, amount) tuples
        receipt_date: Date of the receipt
        file_path: Path to Excel file (defaults to 'receipts.xlsx')

    Raises:
        ProcessingError: If save operation fails
    """
    if file_path is None:
        file_path = settings.excel.default_filename

    # Convert legacy data to Receipt object
    items = []
    for vendor, amount in data:
        item = ReceiptItem(vendor=vendor, amount=Decimal(str(amount)))
        items.append(item)

    receipt = Receipt(date=receipt_date, items=items)

    # Save using new implementation
    writer = ReceiptExcelWriter()
    writer.save_receipt(receipt, Path(file_path))


# Default export manager instance
export_manager = ExcelExportManager()